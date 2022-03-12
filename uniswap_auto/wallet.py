from openpyxl import *


class Excel:

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def getColValues(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows+1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data


def getSeedPhrase(address):
    input_address = address
    # 用户助记词路径，以xlsx格式保存
    file = '/Users/luoye/Downloads/TestNetwork/20220312_eth_uniswap_10.xlsx'
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)

    for i in range(1, len(address_list)):
        if input_address == address_list[i]:
            num = i
            return mnemonic_list[num]
